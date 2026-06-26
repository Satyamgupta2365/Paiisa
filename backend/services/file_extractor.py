"""
File Extractor Service
Handles PDF, Excel (.xlsx/.xls), CSV, and plain text bank statements.
Converts any format to a clean text string for Groq AI to analyse.
"""
import io
import csv


def extract_text_from_file(content: bytes, filename: str) -> str:
    """
    Extracts readable text from any supported bank statement format.
    Supported: .pdf, .xlsx, .xls, .csv, .txt
    Returns a clean string of transaction data.
    """
    name = filename.lower()

    # ── PDF ──────────────────────────────────────────────────────────────────
    if name.endswith(".pdf"):
        return _extract_pdf(content)

    # ── Excel (.xlsx) ─────────────────────────────────────────────────────────
    elif name.endswith(".xlsx"):
        return _extract_xlsx(content)

    # ── Excel (.xls) ──────────────────────────────────────────────────────────
    elif name.endswith(".xls"):
        return _extract_xls(content)

    # ── CSV ───────────────────────────────────────────────────────────────────
    elif name.endswith(".csv"):
        return _extract_csv(content)

    # ── Plain text / any other ────────────────────────────────────────────────
    else:
        try:
            return content.decode("utf-8")
        except UnicodeDecodeError:
            return content.decode("latin-1")


def _extract_pdf(content: bytes) -> str:
    """Extract all text and tables from a PDF bank statement."""
    try:
        import pdfplumber
        lines = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                # Try tables first (most bank statements use tables)
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if row:
                                lines.append("\t".join(str(cell or "") for cell in row))
                else:
                    # Fall back to raw text
                    text = page.extract_text()
                    if text:
                        lines.append(text)
        return "\n".join(lines)
    except Exception as e:
        return f"PDF extraction error: {e}"


def _extract_xlsx(content: bytes) -> str:
    """Extract all sheets from an Excel .xlsx file."""
    try:
        import openpyxl
        lines = []
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    lines.append("\t".join(str(cell) if cell is not None else "" for cell in row))
        return "\n".join(lines)
    except Exception as e:
        return f"Excel extraction error: {e}"


def _extract_xls(content: bytes) -> str:
    """Extract from old-format Excel .xls files."""
    try:
        import xlrd
        lines = []
        wb = xlrd.open_workbook(file_contents=content)
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            lines.append(f"=== Sheet: {sheet_name} ===")
            for row_idx in range(ws.nrows):
                row = ws.row_values(row_idx)
                if any(cell for cell in row):
                    lines.append("\t".join(str(cell) for cell in row))
        return "\n".join(lines)
    except Exception as e:
        return f"XLS extraction error: {e}"


def _extract_csv(content: bytes) -> str:
    """Extract CSV bank statement."""
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    try:
        reader = csv.reader(io.StringIO(text))
        lines = ["\t".join(row) for row in reader if any(cell.strip() for cell in row)]
        return "\n".join(lines)
    except Exception:
        return text
